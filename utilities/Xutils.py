import openpyxl

def getRowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_row)

def getcolumncount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)

def readdata(file,sheetname,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnnum)

def writedata(file, sheetname,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet= workbook[sheetname]
    sheet.cell(row=rownum, column= columnnum).value = data
    workbook.save(file)

